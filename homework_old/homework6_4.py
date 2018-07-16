import json
import csv
from openpyxl import load_workbook


def task_1(path1_json, path2_csv):
    with open(path1_json, 'r') as json_file:
        data = json.load(json_file)
        data_1 = [i for key in data for i in key]
    with open(path2_csv, 'w', newline='') as csv_file:
        fieldnames = set(data_1)
        print(fieldnames)
        writer = csv.DictWriter(csv_file, delimiter=',', fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)


def task_2(path_xls_file, path_json):
    wb = load_workbook(path_xls_file)
    sheet = wb['Лист1']
    array = []
    array_avg = []
    for row in sheet.iter_rows(min_col=3):
        mark = [cell.value for cell in row if cell.value is not None and cell.value != 'end']
        array.append(mark)
    for row in sheet.iter_rows(min_col=3):
        mark = [cell.value for cell in row if cell.value is not None and cell.value != 'end']
        y = sum(mark) / len(mark)
        array_avg.append(y)
    array_of_names = []
    for row in sheet.iter_rows(min_col=1, max_col=2):
        for cell in row:
            if cell.value is not None:
                array_of_names.append(cell.value)
    list1 = array_of_names[::2]
    list2 = array_of_names[1::2]
    sheet.insert_cols(1)
    for i in range(len(list1)):
        sheet.merge_cells('B{}:C{}'.format(i + 1, i + 1))
    i = 1
    while i <= sheet.max_row:
        t = len(array[i - 1])
        g = sheet.cell(i, t + 3)
        sheet['A{}'.format(i)] = '=AVERAGE(D{}:{}{})'.format(i, g.column, g.row)
        i += 1
    for i in range(len(list1)):
        sheet['B' + str(i + 1)].value = list1[i] + ' ' + list2[i]
    wb.save(path_xls_file)
    with open(path_json, 'w', encoding='utf8') as json_file:
        data = {}
        data['students'] = {}
        for i in range(len(array)):
            data1 = {list1[i] + ' ' + list2[i]: {'marks': array[i], 'average_mark': sum(array[i]) / len(array[i])}}
            data['students'].update(data1)
        data['total_average_mark'] = sum(array_avg) / len(array_avg)
        json_file.write(json.dumps(data, ensure_ascii=False, indent=4))
